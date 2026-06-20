"""
Excel Generator – produces the Refund Working Excel matching the sample XLSM structure.
Four sheets:
  1. Post Audit Sheet
  2. MAIN CALCULATION SHEET
  3. Turnover Sheet
  4. ITC Sheet
"""
import io
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter


# ── Colour palette ─────────────────────────────────────────────────────────
DARK_BLUE   = "1F3864"   # header bg
MED_BLUE    = "1F497D"
LIGHT_BLUE  = "BDD7EE"
LIGHT_GREEN = "E2EFDA"
LIGHT_GREY  = "F2F2F2"
ORANGE      = "F4B942"
YELLOW_HDR  = "FFE699"
WHITE       = "FFFFFF"

# ── Style helpers ──────────────────────────────────────────────────────────
def _font(bold=False, size=10, color="000000", name="Calibri"):
    return Font(bold=bold, size=size, color=color, name=name)

def _fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def _align(h="left", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def _thin_border():
    s = Side(style="thin")
    return Border(left=s, right=s, top=s, bottom=s)

def _num_fmt(ws, cell, fmt="#,##0.00"):
    cell.number_format = fmt

def _h(ws, row, col, value, bold=True, bg=LIGHT_BLUE, fg="000000",
       h_align="center", wrap=False, border=True):
    c = ws.cell(row=row, column=col, value=value)
    c.font = _font(bold=bold, color=fg)
    c.fill = _fill(bg)
    c.alignment = _align(h=h_align, v="center", wrap=wrap)
    if border:
        c.border = _thin_border()
    return c

def _d(ws, row, col, value, bold=False, fmt=None, bg=WHITE, h_align="right"):
    c = ws.cell(row=row, column=col, value=value)
    c.font = _font(bold=bold)
    c.fill = _fill(bg)
    c.alignment = _align(h=h_align, v="center")
    c.border = _thin_border()
    if fmt and isinstance(value, (int, float)):
        c.number_format = fmt
    return c

def _fmt_number(v):
    """Round to 2dp for display."""
    if v is None:
        return 0.0
    return round(float(v), 2)


# ── Period helpers ──────────────────────────────────────────────────────────
MONTH_MAP = {
    'january': ('Jan', 1), 'february': ('Feb', 2), 'march': ('Mar', 3),
    'april': ('Apr', 4),   'may': ('May', 5),       'june': ('Jun', 6),
    'july': ('Jul', 7),    'august': ('Aug', 8),    'september': ('Sep', 9),
    'october': ('Oct', 10), 'november': ('Nov', 11), 'december': ('Dec', 12),
}

def _period_label(period_str, year_str):
    """Convert 'January' + '2024-25' → 'Jan-25'."""
    p = str(period_str).strip().lower()
    abbr, mon = MONTH_MAP.get(p, (period_str[:3], 1))
    yr = str(year_str).strip()
    if '-' in yr:
        parts = yr.split('-')
        fy_suffix = parts[1][-2:] if len(parts) > 1 else yr[-2:]
    else:
        fy_suffix = yr[-2:]
    # If month is Apr-Dec → use first year suffix, Jan-Mar → use second year suffix
    if mon >= 4:
        fy_suffix = str(int(parts[0]) % 100) if '-' in yr else yr[-2:]
    return f"{abbr}-{fy_suffix}"


# ═══════════════════════════════════════════════════════════════════════════
# SHEET 1 – Post Audit Sheet
# ═══════════════════════════════════════════════════════════════════════════
def _build_post_audit(ws, g3b, g1, g2b, period_label, net_itc_total, zero_rated, adj_turnover):
    ws.title = "Post Audit Sheet"
    ws.column_dimensions['A'].width = 40
    ws.column_dimensions['B'].width = 14
    ws.column_dimensions['C'].width = 22
    ws.column_dimensions['D'].width = 18
    ws.column_dimensions['E'].width = 22

    # Title
    ws.merge_cells('A1:E1')
    c = ws.cell(row=1, column=1, value="POST AUDIT SHEET")
    c.font = _font(bold=True, size=13, color=WHITE)
    c.fill = _fill(DARK_BLUE)
    c.alignment = _align(h="center", v="center")

    labels_vals = [
        ("NAME", g3b.get('name') or g2b.get('name') or ''),
        ("GSTN NO.", g3b.get('gstin') or g2b.get('gstin') or ''),
        ("ISSUED BY", ''),
        ("DUE DATE OF ACCEPTANCE/REVIEW", ''),
        ("Refund ARN & Date", ''),
        ("RFD-08 (SCN) issued", ''),
        ("If SCN dropped/ confirmed", ''),
        ("Refund Sanction Order No. & date", ''),
        ("Jurisdiction", ''),
        ("REFUND PERIOD", period_label),
        ("GROUND OF REFUND", "Export of Goods - Without payment of Tax"),
    ]
    for i, (lbl, val) in enumerate(labels_vals, start=2):
        ws.merge_cells(f'A{i}:B{i}')
        c = ws.cell(row=i, column=1, value=lbl)
        c.font = _font(bold=True)
        c.fill = _fill(LIGHT_GREY)
        c.alignment = _align()
        c.border = _thin_border()
        ws.merge_cells(f'C{i}:E{i}')
        c2 = ws.cell(row=i, column=3, value=val)
        c2.font = _font()
        c2.alignment = _align()
        c2.border = _thin_border()

    r = len(labels_vals) + 3  # row 14

    ws.merge_cells(f'A{r}:E{r}')
    c = ws.cell(row=r, column=1, value="Amount of Refund Claimed in RFD-01")
    c.font = _font(bold=True, color=WHITE)
    c.fill = _fill(MED_BLUE)
    c.alignment = _align(h="center")

    r += 1
    for col, hdr in enumerate(['IGST', 'CGST', 'SGST', 'Cess', 'Total'], start=1):
        _h(ws, r, col, hdr, bg=YELLOW_HDR, fg="000000")

    r += 1
    igst_claim = _fmt_number(g3b.get('net_itc', {}).get('igst', 0))
    cgst_claim = _fmt_number(g3b.get('net_itc', {}).get('cgst', 0))
    sgst_claim = _fmt_number(g3b.get('net_itc', {}).get('sgst', 0))
    cess_claim = _fmt_number(g3b.get('net_itc', {}).get('cess', 0))
    total_claim = igst_claim + cgst_claim + sgst_claim + cess_claim
    for col, val in enumerate([igst_claim, cgst_claim, sgst_claim, cess_claim, total_claim], start=1):
        _d(ws, r, col, val, fmt="#,##0.00")

    r += 2
    # Statement 3A section
    ws.merge_cells(f'A{r}:E{r}')
    c = ws.cell(row=r, column=1, value="Statement-3A submitted in RFD-01")
    c.font = _font(bold=True, color=WHITE)
    c.fill = _fill(MED_BLUE)
    c.alignment = _align(h="center")

    r += 1
    for col, hdr in enumerate(['', 'Turnover of zero rated supply (1)',
                                'Adjusted total turnover (2)',
                                'Net Input tax credit (3)',
                                'Maximum refund (4)'], start=1):
        _h(ws, r, col, hdr, bg=YELLOW_HDR, wrap=True)

    max_refund = (zero_rated / adj_turnover * net_itc_total) if adj_turnover > 0 else 0.0

    r += 1
    row_vals = [("Integrated Tax", zero_rated, adj_turnover, net_itc_total, max_refund),
                ("Central Tax",    '',         '',           '',            ''),
                ("State/UT Tax",   '',         '',           '',            ''),
                ("CESS",           '',         '',           0.0,           0.0),
                ("Total",          zero_rated, adj_turnover, net_itc_total, max_refund)]

    for label, *vals in row_vals:
        ws.cell(row=r, column=1, value=label).font = _font(bold=(label == "Total"))
        ws.cell(row=r, column=1).alignment = _align()
        ws.cell(row=r, column=1).border = _thin_border()
        for col_off, v in enumerate(vals, start=2):
            c = ws.cell(row=r, column=col_off, value=_fmt_number(v) if isinstance(v, float) else v)
            c.alignment = _align(h="right")
            c.border = _thin_border()
            if isinstance(v, float):
                c.number_format = "#,##0.00"
        r += 1

    r += 1
    ws.merge_cells(f'A{r}:E{r}')
    ws.row_dimensions[r].height = 80
    c = ws.cell(row=r, column=1,
        value="Working of Final Values to be considered for Refund Calculation")
    c.font = _font(bold=True, color=WHITE)
    c.fill = _fill(MED_BLUE)
    c.alignment = _align(h="center")

    r += 1
    for col, hdr in enumerate(['', '(A) Zero rated Supply', 'Domestic Turnover',
                                '(B) Adjusted Total Turnover'], start=1):
        _h(ws, r, col, hdr, bg=YELLOW_HDR, wrap=True)

    domestic_3b = _fmt_number(g3b.get('outward_taxable', {}).get('taxable', 0))
    domestic_1  = (g1.get('b2b_taxable', 0) + g1.get('b2cl_taxable', 0) +
                   g1.get('b2cs_value', 0))
    zero_3b     = _fmt_number(g3b.get('outward_zero_rated', {}).get('taxable', 0))
    zero_1      = _fmt_number(g1.get('exp_wop_value', 0) + g1.get('exp_wp_value', 0) +
                               g1.get('sez_wop_value', 0) + g1.get('sez_wp_value', 0) +
                               g1.get('deemed_export_value', 0))
    adj_3b      = _fmt_number(zero_3b + domestic_3b)
    adj_1       = _fmt_number(zero_1  + domestic_1)

    wfv_rows = [
        ("As per GSTR-3B",     zero_3b, domestic_3b, adj_3b),
        ("As per GSTR-1",      zero_1,  _fmt_number(domestic_1), adj_1),
        ("As per Statement-3A", zero_rated, _fmt_number(domestic_1), adj_turnover),
        ("Final Value",        zero_rated, _fmt_number(domestic_1), adj_turnover),
    ]
    for label, *vals in wfv_rows:
        r += 1
        ws.cell(row=r, column=1, value=label).font = _font(bold=(label == "Final Value"))
        ws.cell(row=r, column=1).alignment = _align()
        ws.cell(row=r, column=1).border = _thin_border()
        for col_off, v in enumerate(vals, start=3):
            c = ws.cell(row=r, column=col_off, value=_fmt_number(v) if isinstance(v, (int, float)) else v)
            c.alignment = _align(h="right")
            c.border = _thin_border()
            if isinstance(v, (int, float)):
                c.number_format = "#,##0.00"

    r += 2
    ws.merge_cells(f'A{r}:E{r}')
    c = ws.cell(row=r, column=1, value="Working of Net ITC value for Refund Calculation")
    c.font = _font(bold=True, color=WHITE)
    c.fill = _fill(MED_BLUE)
    c.alignment = _align(h="center")

    r += 1
    for col, hdr in enumerate(['', '', '', '(IGST+CGST+SGST)', 'CESS'], start=1):
        _h(ws, r, col, hdr, bg=YELLOW_HDR)

    n3b  = _fmt_number(sum([
        g3b.get('net_itc', {}).get('igst', 0),
        g3b.get('net_itc', {}).get('cgst', 0),
        g3b.get('net_itc', {}).get('sgst', 0),
    ]))
    n3b_cess = _fmt_number(g3b.get('net_itc', {}).get('cess', 0))
    n2b  = _fmt_number(g2b.get('net_igst', 0) + g2b.get('net_cgst', 0) + g2b.get('net_sgst', 0))
    n2b_cess = _fmt_number(g2b.get('net_cess', 0))

    nitc_rows = [
        ("Net ITC as per GSTR-3B",     n3b,  n3b_cess),
        ("Net ITC as per GSTR-2B",     n2b,  n2b_cess),
        ("Final Value (min of above)",
         _fmt_number(min(n3b, n2b)), _fmt_number(min(n3b_cess, n2b_cess))),
    ]
    for label, itc_val, cess_val in nitc_rows:
        r += 1
        ws.cell(row=r, column=1, value=label).font = _font(bold=("Final" in label))
        ws.cell(row=r, column=1).alignment = _align()
        ws.cell(row=r, column=1).border = _thin_border()
        for col_off, v in [(4, itc_val), (5, cess_val)]:
            c = ws.cell(row=r, column=col_off, value=v)
            c.alignment = _align(h="right")
            c.border = _thin_border()
            c.number_format = "#,##0.00"

    r += 2
    ws.merge_cells(f'A{r}:E{r}')
    c = ws.cell(row=r, column=1,
        value="Calculation of Maximum Refund Amount as per Rule 89(4)")
    c.font = _font(bold=True, color=WHITE)
    c.fill = _fill(MED_BLUE)
    c.alignment = _align(h="center")

    r += 1
    for col, hdr in enumerate(['', 'Zero Rated (A)', 'Adjusted turnover (B)',
                                'Net ITC (C)', 'Maximum Refund (A/B*C)'], start=1):
        _h(ws, r, col, hdr, bg=YELLOW_HDR, wrap=True)

    r += 1
    igst_max = (zero_rated / adj_turnover * net_itc_total) if adj_turnover > 0 else 0.0
    for label, zr, at, ni, mr in [
        ("Integrated Tax", zero_rated, adj_turnover, net_itc_total, _fmt_number(igst_max)),
        ("Central Tax",    '',         '',            '',            ''),
        ("State/UT Tax",   '',         '',            '',            ''),
        ("CESS",           '',         '',            0.0,           0.0),
        ("Total",          zero_rated, adj_turnover, net_itc_total, _fmt_number(igst_max)),
    ]:
        ws.cell(row=r, column=1, value=label).font = _font(bold=(label == "Total"))
        ws.cell(row=r, column=1).alignment = _align()
        ws.cell(row=r, column=1).border = _thin_border()
        for col_off, v in enumerate([zr, at, ni, mr], start=2):
            c = ws.cell(row=r, column=col_off, value=v)
            c.alignment = _align(h="right")
            c.border = _thin_border()
            if isinstance(v, float):
                c.number_format = "#,##0.00"
        r += 1


# ═══════════════════════════════════════════════════════════════════════════
# SHEET 2 – MAIN CALCULATION SHEET
# ═══════════════════════════════════════════════════════════════════════════
def _build_main_calc(ws, g3b, g1, g2b, period_label, net_itc_total, zero_rated, adj_turnover):
    ws.title = "MAIN CALCULATION SHEET"
    ws.freeze_panes = 'B2'

    hdr_cols = ['A','B','C','D','E','F','G','H','I','J','K','L','M','N','O','P','Q','R','S','T','U','V','W','X','Y','Z']
    col_widths = [6, 22, 18, 28, 26, 18, 20, 30, 16, 20, 24, 10, 28, 24, 16, 14, 14, 14, 14, 14, 14, 14, 14, 14, 24, 24]
    for i, (letter, w) in enumerate(zip(hdr_cols, col_widths)):
        ws.column_dimensions[letter].width = w

    # Row 1 – Main header
    headers = [
        'Sr.No', 'DD Division / Range', 'GSTIN', 'Name of the Taxpayer',
        'ARN & DATE', 'Amount in ARN (Rs.)', 'Refund Period',
        'RFD06 Order No. and Date', 'Amount sanctioned',
        'Post Audit Objections if any',
        'Is the refund claim found to be in order (Yes/No)',
        'Division', 'Refund Category', 'Officer Name',
        'Due date for review', 'Legal Name',
        'CGST-Claimed', 'SGST-Claimed', 'IGST-Claimed', 'Cess-Claimed',
        'CGST-sanctioned', 'SGST-sanctioned', 'IGST-sanctioned', 'Cess-sanctioned',
        'RFD-08 & date', 'SCN dropped/confirmed',
    ]
    for col, hdr in enumerate(headers, start=1):
        _h(ws, 1, col, hdr, bg=DARK_BLUE, fg=WHITE, wrap=True)

    # Row 2 – Data
    igst  = _fmt_number(g3b.get('net_itc', {}).get('igst', 0))
    cgst  = _fmt_number(g3b.get('net_itc', {}).get('cgst', 0))
    sgst  = _fmt_number(g3b.get('net_itc', {}).get('sgst', 0))
    cess  = _fmt_number(g3b.get('net_itc', {}).get('cess', 0))
    total = igst + cgst + sgst + cess
    max_r = _fmt_number((zero_rated / adj_turnover * net_itc_total) if adj_turnover > 0 else 0)

    data_row = [
        '',
        '',
        g3b.get('gstin') or g2b.get('gstin', ''),
        g3b.get('name') or g2b.get('name', ''),
        '',
        total,
        period_label,
        '', '', '', '', '', 'Export of Goods - Without payment of Tax',
        '', '', g3b.get('name') or g2b.get('name', ''),
        cgst, sgst, igst, cess,
        '', '', '', '', '', '',
    ]
    for col, val in enumerate(data_row, start=1):
        c = ws.cell(row=2, column=col, value=val)
        c.alignment = _align(h="left" if isinstance(val, str) else "right")
        c.border = _thin_border()
        if isinstance(val, float):
            c.number_format = "#,##0.00"

    # ── Section A – Refund claimed ─────────────────────────────────────────
    ws.cell(row=5, column=2, value="A) Amount of Refund Claimed").font = _font(bold=True)
    for col, hdr in enumerate(['IGST', 'CGST', 'SGST', 'Cess', 'Total'], start=2):
        _h(ws, 6, col, hdr, bg=YELLOW_HDR)
    for col, val in enumerate([igst, cgst, sgst, cess, total], start=2):
        _d(ws, 7, col, val, fmt="#,##0.00")

    # ── Section B – Statement 3A ───────────────────────────────────────────
    ws.cell(row=14, column=2, value="B) Statement-3A").font = _font(bold=True)
    for col, hdr in enumerate(['', 'Turnover of zero rated supply (1)',
                                'Adjusted total turnover (2)',
                                'Net Input tax credit (3)',
                                'Maximum refund (4)'], start=2):
        _h(ws, 15, col, hdr, bg=YELLOW_HDR, wrap=True)

    max_r2 = _fmt_number((zero_rated / adj_turnover * net_itc_total) if adj_turnover > 0 else 0)
    s3a_rows = [
        ("Integrated Tax", zero_rated, adj_turnover, net_itc_total, max_r2),
        ("Central Tax",    '',         '',            '',            ''),
        ("State/UT Tax",   '',         '',            '',            ''),
        ("CESS",           '',         '',            0.0,           0.0),
        ("Total",          zero_rated, adj_turnover, net_itc_total, max_r2),
    ]
    for i, (label, *vals) in enumerate(s3a_rows, start=16):
        ws.cell(row=i, column=2, value=label).font = _font(bold=(label == "Total"))
        ws.cell(row=i, column=2).border = _thin_border()
        for j, v in enumerate(vals, start=3):
            c = ws.cell(row=i, column=j, value=v)
            c.alignment = _align(h="right")
            c.border = _thin_border()
            if isinstance(v, float):
                c.number_format = "#,##0.00"

    # ── Section C – Working of final values ───────────────────────────────
    ws.cell(row=5, column=9, value="C) Working of Final Values").font = _font(bold=True)
    for col, hdr in enumerate(['', 'Zero rated Supply', 'Domestic Turnover',
                                'Adjusted Total Turnover'], start=9):
        _h(ws, 6, col, hdr, bg=YELLOW_HDR, wrap=True)

    domestic_3b = _fmt_number(g3b.get('outward_taxable', {}).get('taxable', 0))
    nil_3b      = _fmt_number(g3b.get('outward_nil_exempted', {}).get('taxable', 0))
    zero_3b     = _fmt_number(g3b.get('outward_zero_rated', {}).get('taxable', 0))
    zero_1      = _fmt_number(g1.get('exp_wop_value', 0) + g1.get('exp_wp_value', 0))
    domestic_1  = _fmt_number(g1.get('b2b_taxable', 0) + g1.get('b2cl_taxable', 0) +
                               g1.get('b2cs_value', 0))
    adj_3b      = _fmt_number(zero_3b + domestic_3b)
    adj_1       = _fmt_number(zero_1  + domestic_1)

    c_rows = [
        ("As per GSTR-3B",  zero_3b, domestic_3b, adj_3b),
        ("As per GSTR-1",   zero_1,  domestic_1,  adj_1),
        ("As per Stmt-3A",  zero_rated, domestic_1, adj_turnover),
        ("Final Value",     zero_rated, domestic_1, adj_turnover),
    ]
    for i, (label, *vals) in enumerate(c_rows, start=7):
        ws.cell(row=i, column=9, value=label).font = _font(bold=(label == "Final Value"))
        ws.cell(row=i, column=9).border = _thin_border()
        for j, v in enumerate(vals, start=10):
            c = ws.cell(row=i, column=j, value=v)
            c.alignment = _align(h="right")
            c.border = _thin_border()
            if isinstance(v, float):
                c.number_format = "#,##0.00"

    # ── ITC comparison table ───────────────────────────────────────────────
    ws.cell(row=15, column=9, value="Working of Net ITC").font = _font(bold=True)
    for col, hdr in enumerate(['', '', 'GSTR-3B', 'GSTR-2B'], start=9):
        _h(ws, 16, col, hdr, bg=YELLOW_HDR)

    n3b_all_other = _fmt_number(
        g3b.get('itc_all_other', {}).get('igst', 0) +
        g3b.get('itc_all_other', {}).get('cgst', 0) +
        g3b.get('itc_all_other', {}).get('sgst', 0))
    n2b_all_other = _fmt_number(g2b.get('4a5_igst', 0) + g2b.get('4a5_cgst', 0) + g2b.get('4a5_sgst', 0))

    itc_comparison = [
        ("ITC under 'All Other ITC'",      n3b_all_other,  n2b_all_other),
        ("ITC under 'Import'",
         _fmt_number(g3b.get('itc_import_goods', {}).get('igst', 0)),
         _fmt_number(g2b.get('import_igst', 0))),
        ("ITC under 'RCM'",
         _fmt_number(g3b.get('itc_rcm', {}).get('igst', 0) + g3b.get('itc_rcm', {}).get('cgst', 0) + g3b.get('itc_rcm', {}).get('sgst', 0)),
         0.0),
        ("ITC under 'ISD'",
         _fmt_number(g3b.get('itc_isd', {}).get('igst', 0) + g3b.get('itc_isd', {}).get('cgst', 0) + g3b.get('itc_isd', {}).get('sgst', 0)),
         _fmt_number(g2b.get('4a4_igst', 0) + g2b.get('4a4_cgst', 0) + g2b.get('4a4_sgst', 0))),
        ("Total of Cr. & Db. Notes",
         0.0,
         _fmt_number(g2b.get('cdn_igst', 0) + g2b.get('cdn_cgst', 0) + g2b.get('cdn_sgst', 0))),
        ("ITC Reversed",
         _fmt_number(g3b.get('itc_reversed_2', {}).get('igst', 0) + g3b.get('itc_reversed_2', {}).get('cgst', 0) + g3b.get('itc_reversed_2', {}).get('sgst', 0)),
         0.0),
        ("ITC for Claimed Tax Period",
         _fmt_number(sum(g3b.get('net_itc', {}).values())),
         _fmt_number(g2b.get('net_igst', 0) + g2b.get('net_cgst', 0) + g2b.get('net_sgst', 0))),
    ]
    for i, (label, v3b, v2b) in enumerate(itc_comparison, start=17):
        ws.cell(row=i, column=9, value=label).font = _font(bold=(label == "ITC for Claimed Tax Period"))
        ws.cell(row=i, column=9).border = _thin_border()
        for j, v in [(11, v3b), (12, v2b)]:
            c = ws.cell(row=i, column=j, value=v)
            c.alignment = _align(h="right")
            c.border = _thin_border()
            if isinstance(v, float):
                c.number_format = "#,##0.00"

    # ── Maximum refund calc ───────────────────────────────────────────────
    ws.cell(row=22, column=2, value="Calculation of Maximum Refund Amount as per Rule 89(4)").font = _font(bold=True)
    for col, hdr in enumerate(['', 'Zero Rated', 'Adjusted turnover', 'Net ITC', 'Maximum Refund'], start=2):
        _h(ws, 23, col, hdr, bg=YELLOW_HDR, wrap=True)
    for i, (label, zr, at, ni, mr) in enumerate([
        ("Integrated Tax", zero_rated, adj_turnover, net_itc_total,
         _fmt_number((zero_rated / adj_turnover * net_itc_total) if adj_turnover > 0 else 0)),
        ("Central Tax",    '', '', '', ''),
        ("State/UT Tax",   '', '', '', ''),
        ("CESS",           '', '', 0.0, 0.0),
        ("Total", zero_rated, adj_turnover, net_itc_total,
         _fmt_number((zero_rated / adj_turnover * net_itc_total) if adj_turnover > 0 else 0)),
    ], start=24):
        ws.cell(row=i, column=2, value=label).font = _font(bold=(label == "Total"))
        ws.cell(row=i, column=2).border = _thin_border()
        for j, v in enumerate([zr, at, ni, mr], start=3):
            c = ws.cell(row=i, column=j, value=v)
            c.alignment = _align(h="right")
            c.border = _thin_border()
            if isinstance(v, float):
                c.number_format = "#,##0.00"


# ═══════════════════════════════════════════════════════════════════════════
# SHEET 3 – Turnover Sheet
# ═══════════════════════════════════════════════════════════════════════════
def _build_turnover(ws, g3b, g1, period_label):
    ws.title = "Turnover Sheet"
    ws.freeze_panes = 'B7'

    # Column widths
    ws.column_dimensions['A'].width = 10
    for col in range(2, 35):
        ws.column_dimensions[get_column_letter(col)].width = 14

    # Row 1 – Period header
    ws.cell(row=1, column=2, value="From").font = _font(bold=True)
    ws.cell(row=1, column=3, value="To").font = _font(bold=True)
    ws.cell(row=2, column=1, value="Tax period for Refund Claim").font = _font(bold=True)
    ws.cell(row=2, column=2, value=period_label)
    ws.cell(row=2, column=3, value=period_label)

    # Row 4 – Section headers
    ws.merge_cells('A4:E4')
    c = ws.cell(row=4, column=1, value="Turnover as per GSTR-3B")
    c.font = _font(bold=True, color=WHITE)
    c.fill = _fill(DARK_BLUE)
    c.alignment = _align(h="center")

    ws.merge_cells('F4:AD4')
    c = ws.cell(row=4, column=6, value="Turnover as per GSTR-1")
    c.font = _font(bold=True, color=WHITE)
    c.fill = _fill(MED_BLUE)
    c.alignment = _align(h="center")

    # Row 5 – Sub-headers
    col5 = [
        (1, "Period"), (2, "Domestic Sales"), (3, "Zero rated"), (4, "(nil rated, exempted)"),
    ]
    for col, hdr in col5:
        _h(ws, 5, col, hdr, bg=LIGHT_BLUE, wrap=True)

    gstr1_sub = [
        (6, "Period"), (7, "DOMESTIC SALES"), (13, ""), (14, ""), (15, ""), (16, ""), (17, ""),
        (18, "DOMESTIC TOTAL"), (19, "ZERO RATED SUPPLIES"),
        (25, ""), (26, ""),
        (30, "TOTAL OF ZERO RATED SUPPLIES"),
    ]
    for col, hdr in gstr1_sub:
        _h(ws, 5, col, hdr, bg=LIGHT_GREEN if col >= 6 else LIGHT_BLUE, wrap=True)

    # Row 6 – Detailed sub-headers
    gstr3b_r6 = [(2, "3.1(a)"), (3, "3.1 (b)"), (4, "3.1 (c)")]
    for col, hdr in gstr3b_r6:
        _h(ws, 6, col, hdr, bg=LIGHT_BLUE, wrap=True)

    gstr1_r6 = [
        (7, "B2B (4A)"), (8, "B2CL (5)"), (9, "B2CS (7)"),
        (10, "CDNR (9B)\ni.r.o. (4A)"), (11, "CDNUR (9B)\ni.r.o. B2CL"),
        (12, "AMENDMENTS (9A)"), (13, "ADVANCE (11A)"), (14, "ADV ADJS (11B)"),
        (15, "NIL rated"), (16, "Exempted"), (17, "Non-GST"),
        (18, "Total (Domestic)"),
        (19, "EXP WOP (6A)"), (20, "EXP WP (6A)"),
        (21, "SEZ WOP (6B)"), (22, "SEZ WP (6B)"),
        (23, "Deemed Export (6C)"),
        (24, "CDNUR EXP WOP\n(9B)"), (25, "CDNUR EXP WP\n(9B)"),
        (26, "CDNR SEZ WOP"), (27, "CDNR Deemed Exp"),
        (28, "ZRS Turnover\n(EXP WOP)"), (29, "ZRS Turnover\n(EXP WP)"),
        (30, "Zero rated\n(Export)"), (31, "Zero rated\n(SEZ)"), (32, "Zero rated\n(Deemed)"),
    ]
    for col, hdr in gstr1_r6:
        _h(ws, 6, col, hdr, bg=LIGHT_GREEN, wrap=True, h_align="center")
    ws.row_dimensions[6].height = 42

    # ── Data row ──────────────────────────────────────────────────────────
    row = 7
    domestic_3b = _fmt_number(g3b.get('outward_taxable', {}).get('taxable', 0))
    zero_3b     = _fmt_number(g3b.get('outward_zero_rated', {}).get('taxable', 0))
    nil_3b      = _fmt_number(g3b.get('outward_nil_exempted', {}).get('taxable', 0))

    b2b   = _fmt_number(g1.get('b2b_taxable', 0))
    b2cl  = _fmt_number(g1.get('b2cl_taxable', 0))
    b2cs  = _fmt_number(g1.get('b2cs_value', 0))
    cdnr  = _fmt_number(-g1.get('cdnr_value', 0))   # CDN reduces
    nil1  = _fmt_number(g1.get('nil_rated', 0))
    exm1  = _fmt_number(g1.get('exempted', 0))
    ngst1 = _fmt_number(g1.get('non_gst', 0))
    dom_total = _fmt_number(b2b + b2cl + b2cs + cdnr)

    exp_wop = _fmt_number(g1.get('exp_wop_value', 0))
    exp_wp  = _fmt_number(g1.get('exp_wp_value', 0))
    sez_wop = _fmt_number(g1.get('sez_wop_value', 0))
    sez_wp  = _fmt_number(g1.get('sez_wp_value', 0))
    deemed  = _fmt_number(g1.get('deemed_export_value', 0))
    zrs_exp_wop = exp_wop   # ZRS = EXP WOP value
    zrs_total   = _fmt_number(exp_wop + exp_wp + sez_wop + sez_wp + deemed)

    data_3b = {1: period_label, 2: domestic_3b, 3: zero_3b, 4: nil_3b}
    data_g1 = {
        7: b2b, 8: b2cl, 9: b2cs, 10: cdnr, 11: 0.0, 12: 0.0, 13: 0.0, 14: 0.0,
        15: nil1, 16: exm1, 17: ngst1, 18: dom_total,
        19: exp_wop, 20: exp_wp, 21: sez_wop, 22: sez_wp, 23: deemed,
        24: 0.0, 25: 0.0, 26: 0.0, 27: 0.0,
        28: zrs_exp_wop, 29: exp_wp,
        30: zrs_total, 31: _fmt_number(sez_wop + sez_wp), 32: deemed,
    }
    for col, val in {**data_3b, **data_g1}.items():
        c = ws.cell(row=row, column=col, value=val)
        c.border = _thin_border()
        c.alignment = _align(h="right" if isinstance(val, float) else "center")
        if isinstance(val, float):
            c.number_format = "#,##0.00"

    # Total row
    row += 1
    ws.cell(row=row, column=1, value="Total").font = _font(bold=True)
    ws.cell(row=row, column=1).fill = _fill(LIGHT_GREY)
    ws.cell(row=row, column=1).border = _thin_border()
    total_map = {
        2: domestic_3b, 3: zero_3b, 4: nil_3b,
        7: b2b, 8: b2cl, 9: b2cs, 10: cdnr, 15: nil1, 16: exm1, 17: ngst1,
        18: dom_total, 19: exp_wop, 20: exp_wp, 21: sez_wop, 22: sez_wp,
        23: deemed, 28: zrs_exp_wop, 29: exp_wp,
        30: zrs_total, 31: _fmt_number(sez_wop + sez_wp), 32: deemed,
    }
    for col, val in total_map.items():
        c = ws.cell(row=row, column=col, value=val)
        c.font = _font(bold=True)
        c.fill = _fill(LIGHT_GREY)
        c.border = _thin_border()
        c.alignment = _align(h="right")
        c.number_format = "#,##0.00"

    # Summary labels below
    row += 2
    for label, val in [("DOMESTIC GSTR-3B", domestic_3b), ("Export GSTR-3B", zero_3b)]:
        ws.cell(row=row, column=1, value=label).font = _font(bold=True)
        ws.cell(row=row, column=2, value=val).number_format = "#,##0.00"
        row += 1


# ═══════════════════════════════════════════════════════════════════════════
# SHEET 4 – ITC Sheet
# ═══════════════════════════════════════════════════════════════════════════
def _build_itc(ws, g3b, g2b, period_label):
    ws.title = "ITC Sheet"
    ws.freeze_panes = 'B7'

    ws.column_dimensions['A'].width = 10
    for col in range(2, 38):
        ws.column_dimensions[get_column_letter(col)].width = 11

    # Row 1-2 header
    ws.cell(row=1, column=2, value="From").font = _font(bold=True)
    ws.cell(row=1, column=3, value="To").font = _font(bold=True)
    ws.cell(row=2, column=1, value="Tax period for Refund Claim").font = _font(bold=True)
    ws.cell(row=2, column=2, value=period_label)
    ws.cell(row=2, column=3, value=period_label)

    # Row 4 – Section title
    ws.merge_cells('A4:AK4')
    c = ws.cell(row=4, column=1, value="ITC availed as per GSTR-3B")
    c.font = _font(bold=True, color=WHITE)
    c.fill = _fill(DARK_BLUE)
    c.alignment = _align(h="center")

    # Row 5 – Section group headers
    GSTR3B_COLS = [
        (2,  "4. (A) (1)\nImport of Goods",                1, 'B'),
        (3,  "4. (A) (2)\nImport of Services",              1, 'C'),
        (4,  "4. (A) (3)\nRCM (other than 1 & 2)",          4, 'D'),
        (8,  "4. (A) (4)\nInward supplies from ISD",         4, 'H'),
        (12, "4. (A) (5)\nAll other ITC",                   4, 'L'),
        (16, "4. (A) NET TOTAL",                             4, 'P'),
        (20, "4. (B)(1) ITC Reversed\n(Rules 38,42,43)",    4, 'T'),
        (24, "4. (B)(2) ITC Reversed\nOthers",              4, 'X'),
        (28, "4. (C) Net ITC Available",                    4, 'AB'),
        (32, "ITC RECLAIMED (4D)",                          4, 'AF'),
    ]
    for start_col, hdr, span, _ in GSTR3B_COLS:
        if span > 1:
            ws.merge_cells(start_row=5, start_column=start_col,
                           end_row=5, end_column=start_col + span - 1)
        _h(ws, 5, start_col, hdr, bg=LIGHT_BLUE, wrap=True, h_align="center")
        ws.row_dimensions[5].height = 40

    # Row 6 – IGST/CGST/SGST/CESS sub-headers
    ws.cell(row=6, column=1, value='').border = _thin_border()
    ws.cell(row=6, column=2, value='IGST').font = _font(bold=True)
    ws.cell(row=6, column=2).border = _thin_border()
    ws.cell(row=6, column=2).fill = _fill(LIGHT_BLUE)
    ws.cell(row=6, column=3, value='IGST').font = _font(bold=True)
    ws.cell(row=6, column=3).border = _thin_border()
    ws.cell(row=6, column=3).fill = _fill(LIGHT_BLUE)

    sub_labels = ['IGST', 'CGST', 'SGST', 'CESS']
    for start_col, _, span, _ in GSTR3B_COLS:
        if span == 4:
            for i, lbl in enumerate(sub_labels):
                c = ws.cell(row=6, column=start_col + i, value=lbl)
                c.font = _font(bold=True)
                c.fill = _fill(LIGHT_BLUE)
                c.alignment = _align(h="center")
                c.border = _thin_border()

    # ── GSTR-3B Data row ──────────────────────────────────────────────────
    row = 7
    g = g3b
    def igst(d): return _fmt_number(d.get('igst', 0))
    def cgst(d): return _fmt_number(d.get('cgst', 0))
    def sgst(d): return _fmt_number(d.get('sgst', 0))
    def cess(d): return _fmt_number(d.get('cess', 0))

    rcm    = g.get('itc_rcm', {})
    isd    = g.get('itc_isd', {})
    ao     = g.get('itc_all_other', {})
    rev1   = g.get('itc_reversed_1', {})
    rev2   = g.get('itc_reversed_2', {})
    nitc   = g.get('net_itc', {})
    reclm  = g.get('itc_reclaimed', {})
    imp_g  = g.get('itc_import_goods', {})
    imp_s  = g.get('itc_import_svc', {})

    net_p_igst = _fmt_number(igst(imp_g) + igst(imp_s) + igst(rcm) + igst(isd) + igst(ao))
    net_p_cgst = _fmt_number(cgst(rcm) + cgst(isd) + cgst(ao))
    net_p_sgst = _fmt_number(sgst(rcm) + sgst(isd) + sgst(ao))
    net_p_cess = _fmt_number(cess(rcm) + cess(isd) + cess(ao))

    col_vals = {
        1: period_label,
        2: igst(imp_g),
        3: igst(imp_s),
        4: igst(rcm), 5: cgst(rcm), 6: sgst(rcm), 7: cess(rcm),
        8: igst(isd), 9: cgst(isd), 10: sgst(isd), 11: cess(isd),
        12: igst(ao), 13: cgst(ao), 14: sgst(ao), 15: cess(ao),
        16: net_p_igst, 17: net_p_cgst, 18: net_p_sgst, 19: net_p_cess,
        20: igst(rev1), 21: cgst(rev1), 22: sgst(rev1), 23: cess(rev1),
        24: igst(rev2), 25: cgst(rev2), 26: sgst(rev2), 27: cess(rev2),
        28: igst(nitc), 29: cgst(nitc), 30: sgst(nitc), 31: cess(nitc),
        32: igst(reclm), 33: cgst(reclm), 34: sgst(reclm), 35: cess(reclm),
    }
    for col, val in col_vals.items():
        c = ws.cell(row=row, column=col, value=val)
        c.border = _thin_border()
        c.alignment = _align(h="right" if isinstance(val, float) else "center")
        if isinstance(val, float):
            c.number_format = "#,##0.00"

    # Total row (GSTR-3B section)
    row += 1  # skip some blank rows, then total
    for i in range(2):
        row2 = row + i
        for col in range(1, 36):
            ws.cell(row=row2, column=col).border = _thin_border()

    row += 2
    ws.cell(row=row, column=1, value="Total").font = _font(bold=True)
    ws.cell(row=row, column=1).fill = _fill(LIGHT_GREY)
    ws.cell(row=row, column=1).border = _thin_border()
    for col, val in col_vals.items():
        if col == 1:
            continue
        c = ws.cell(row=row, column=col, value=val)
        c.font = _font(bold=True)
        c.fill = _fill(LIGHT_GREY)
        c.border = _thin_border()
        c.alignment = _align(h="right")
        if isinstance(val, float):
            c.number_format = "#,##0.00"

    # ── GSTR-2B Section ───────────────────────────────────────────────────
    row += 3
    ws.merge_cells(f'A{row}:AK{row}')
    c = ws.cell(row=row, column=1, value="ITC availed as per GSTR-2B")
    c.font = _font(bold=True, color=WHITE)
    c.fill = _fill(MED_BLUE)
    c.alignment = _align(h="center")

    row += 1
    # Sub-section headers for GSTR-2B
    G2B_COLS = [
        (2,  "4. (A) (1)\nImport of Goods (IGST)", 1),
        (3,  "4. (A) (2)\nImport Services\n(from GSTR-3B)", 1),
        (8,  "4. (A) (4)\nISD",                   4),
        (12, "4. (A) (5)\nAll other ITC",          4),
        (16, "4. (A) NET TOTAL",                    4),
        (24, "Debit & Credit Notes (B)",            4),
        (28, "4. (C) Net ITC Available",            4),
    ]
    for start_col, hdr, span in G2B_COLS:
        if span > 1:
            ws.merge_cells(start_row=row, start_column=start_col,
                           end_row=row, end_column=start_col + span - 1)
        _h(ws, row, start_col, hdr, bg=LIGHT_GREEN, wrap=True, h_align="center")
    ws.row_dimensions[row].height = 42

    row += 1
    ws.cell(row=row, column=2, value='IGST').fill = _fill(LIGHT_GREEN)
    ws.cell(row=row, column=2).border = _thin_border()
    ws.cell(row=row, column=2).font = _font(bold=True)
    ws.cell(row=row, column=3, value='IGST').fill = _fill(LIGHT_GREEN)
    ws.cell(row=row, column=3).border = _thin_border()
    ws.cell(row=row, column=3).font = _font(bold=True)
    for start_col, _, span in G2B_COLS:
        if span == 4:
            for i, lbl in enumerate(sub_labels):
                c2 = ws.cell(row=row, column=start_col + i, value=lbl)
                c2.font = _font(bold=True)
                c2.fill = _fill(LIGHT_GREEN)
                c2.alignment = _align(h="center")
                c2.border = _thin_border()

    # GSTR-2B data row
    row += 1
    cdn_net_igst = _fmt_number(g2b.get('cdn_igst', 0))  # negative = credit note
    cdn_net_cgst = _fmt_number(g2b.get('cdn_cgst', 0))
    cdn_net_sgst = _fmt_number(g2b.get('cdn_sgst', 0))
    cdn_net_cess = _fmt_number(g2b.get('cdn_cess', 0))

    g2b_4a5_igst = _fmt_number(g2b.get('4a5_igst', 0))
    g2b_4a5_cgst = _fmt_number(g2b.get('4a5_cgst', 0))
    g2b_4a5_sgst = _fmt_number(g2b.get('4a5_sgst', 0))
    g2b_4a5_cess = _fmt_number(g2b.get('4a5_cess', 0))

    imp_igst = _fmt_number(g2b.get('import_igst', 0))
    g2b_isd_igst = _fmt_number(g2b.get('4a4_igst', 0))
    g2b_isd_cgst = _fmt_number(g2b.get('4a4_cgst', 0))
    g2b_isd_sgst = _fmt_number(g2b.get('4a4_sgst', 0))
    g2b_isd_cess = _fmt_number(g2b.get('4a4_cess', 0))

    net2b_igst = _fmt_number(g2b.get('net_igst', 0))
    net2b_cgst = _fmt_number(g2b.get('net_cgst', 0))
    net2b_sgst = _fmt_number(g2b.get('net_sgst', 0))
    net2b_cess = _fmt_number(g2b.get('net_cess', 0))

    net_total_igst = _fmt_number(imp_igst + g2b_isd_igst + g2b_4a5_igst)
    net_total_cgst = _fmt_number(g2b_isd_cgst + g2b_4a5_cgst)
    net_total_sgst = _fmt_number(g2b_isd_sgst + g2b_4a5_sgst)
    net_total_cess = _fmt_number(g2b_isd_cess + g2b_4a5_cess)

    g2b_col_vals = {
        1: period_label,
        2: imp_igst, 3: 0.0,
        8: g2b_isd_igst, 9: g2b_isd_cgst, 10: g2b_isd_sgst, 11: g2b_isd_cess,
        12: g2b_4a5_igst, 13: g2b_4a5_cgst, 14: g2b_4a5_sgst, 15: g2b_4a5_cess,
        16: net_total_igst, 17: net_total_cgst, 18: net_total_sgst, 19: net_total_cess,
        24: cdn_net_igst, 25: cdn_net_cgst, 26: cdn_net_sgst, 27: cdn_net_cess,
        28: net2b_igst, 29: net2b_cgst, 30: net2b_sgst, 31: net2b_cess,
    }
    for col, val in g2b_col_vals.items():
        c = ws.cell(row=row, column=col, value=val)
        c.border = _thin_border()
        c.alignment = _align(h="right" if isinstance(val, float) else "center")
        if isinstance(val, float):
            c.number_format = "#,##0.00"

    # Total row (GSTR-2B)
    row += 3
    ws.cell(row=row, column=1, value="Total").font = _font(bold=True)
    ws.cell(row=row, column=1).fill = _fill(LIGHT_GREY)
    ws.cell(row=row, column=1).border = _thin_border()
    for col, val in g2b_col_vals.items():
        if col == 1:
            continue
        c = ws.cell(row=row, column=col, value=val)
        c.font = _font(bold=True)
        c.fill = _fill(LIGHT_GREY)
        c.border = _thin_border()
        c.alignment = _align(h="right")
        if isinstance(val, float):
            c.number_format = "#,##0.00"

    # ── Annexure-B ────────────────────────────────────────────────────────
    row += 4
    ws.merge_cells(f'N{row}:R{row}')
    c = ws.cell(row=row, column=14, value="ANNEXURE-B")
    c.font = _font(bold=True, color=WHITE)
    c.fill = _fill(DARK_BLUE)
    c.alignment = _align(h="center")

    row += 1
    for col, hdr in enumerate(['', 'IGST', 'CGST', 'SGST', 'CESS', 'Total'], start=13):
        _h(ws, row, col, hdr, bg=YELLOW_HDR)

    # Annexure totals
    annexb_rows = [
        ("Total eligible",
         _fmt_number(igst(nitc) + cgst(nitc) + sgst(nitc) + cess(nitc)),
         _fmt_number(igst(nitc)), _fmt_number(cgst(nitc)), _fmt_number(sgst(nitc)), _fmt_number(cess(nitc))),
        ("Not available", 0.0, 0.0, 0.0, 0.0, 0.0),
        ("Net ITC for refund",
         _fmt_number(igst(nitc) + cgst(nitc) + sgst(nitc) + cess(nitc)),
         _fmt_number(igst(nitc)), _fmt_number(cgst(nitc)), _fmt_number(sgst(nitc)), _fmt_number(cess(nitc))),
    ]
    for label, total, igst_v, cgst_v, sgst_v, cess_v in annexb_rows:
        row += 1
        ws.cell(row=row, column=13, value=label).font = _font(bold=("Net ITC" in label))
        ws.cell(row=row, column=13).border = _thin_border()
        for col, v in enumerate([igst_v, cgst_v, sgst_v, cess_v, total], start=14):
            c = ws.cell(row=row, column=col, value=v)
            c.alignment = _align(h="right")
            c.border = _thin_border()
            c.number_format = "#,##0.00"


# ═══════════════════════════════════════════════════════════════════════════
# MAIN ENTRY POINT
# ═══════════════════════════════════════════════════════════════════════════
def generate_refund_excel(g3b, g1, g2b):
    """
    Generate the Refund Working Excel.
    Returns bytes of the .xlsx file.
    """
    # ── Derived values ────────────────────────────────────────────────────
    period_str = g3b.get('period') or g2b.get('period') or 'Unknown'
    year_str   = g3b.get('year')   or g2b.get('year')   or '2024-25'
    period_label = _period_label(period_str, year_str)

    # Net ITC total (IGST + CGST + SGST + CESS from GSTR-3B)
    nitc = g3b.get('net_itc', {})
    net_itc_total = _fmt_number(
        nitc.get('igst', 0) + nitc.get('cgst', 0) + nitc.get('sgst', 0) + nitc.get('cess', 0))

    # Zero-rated turnover (Export WOP from GSTR-1)
    zero_rated = _fmt_number(
        g1.get('exp_wop_value', 0) + g1.get('exp_wp_value', 0) +
        g1.get('sez_wop_value', 0) + g1.get('sez_wp_value', 0) +
        g1.get('deemed_export_value', 0))

    # Adjusted total turnover = zero rated + domestic
    domestic = _fmt_number(
        g3b.get('outward_taxable', {}).get('taxable', 0) +
        g3b.get('outward_nil_exempted', {}).get('taxable', 0))
    adj_turnover = _fmt_number(zero_rated + domestic)

    # ── Workbook ──────────────────────────────────────────────────────────
    wb = Workbook()
    # Remove default sheet
    wb.remove(wb.active)

    ws_post   = wb.create_sheet("Post Audit Sheet")
    ws_main   = wb.create_sheet("MAIN CALCULATION SHEET")
    ws_turn   = wb.create_sheet("Turnover Sheet")
    ws_itc    = wb.create_sheet("ITC Sheet")

    _build_post_audit(ws_post,  g3b, g1, g2b, period_label, net_itc_total, zero_rated, adj_turnover)
    _build_main_calc( ws_main,  g3b, g1, g2b, period_label, net_itc_total, zero_rated, adj_turnover)
    _build_turnover(  ws_turn,  g3b, g1, period_label)
    _build_itc(       ws_itc,   g3b, g2b, period_label)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()
